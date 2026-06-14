"""Excel 多 sheet 输出(可选依赖,走 ``[post]`` extras)。

依赖:``openpyxl``。顶层不 import — 调用 ``write_excel`` 时按需 lazy import。
未装时抛 ``ImportError`` 提示安装。

参考实现:reference/code/CFDPlus_V4.py:742-793 + CFDPlus_extract.py:449-495

设计要点:
- ``write_excel(force_report, out_path, sheet_name=None) -> Path``
- 默认 sheet 名 "Forces",可覆盖;> 31 字符自动截断(Excel 限制)
- 28 列 header,字段顺序与 ``CoefficientRow`` 一致
- header 字体 Times New Roman 10pt 粗体居中;data 行常规右对齐
- 列宽自适应:``min(max_len + 3, 25)``
- 空 ``ForceReport`` 只写 header 行
"""
from __future__ import annotations

from dataclasses import fields as dataclass_fields
from pathlib import Path
from typing import Optional, Union

from .forces import CoefficientRow, ForceReport

# Excel sheet 名长度上限
_SHEET_NAME_MAX_LEN = 31
_DEFAULT_SHEET_NAME = "Forces"

# 字体规格
_FONT_NAME = "Times New Roman"
_FONT_SIZE_HEADER = 10
_FONT_SIZE_DATA = 10

# 列宽计算
_COL_WIDTH_PADDING = 3
_COL_WIDTH_MAX = 25

# 28 列 header(对应 CoefficientRow 字段顺序;含单位标注便于阅读)
HEADERS = [
    "Case",
    "Ma",
    "H(km)",
    "Alpha(deg)",
    "Beta(deg)",
    "Fx(N)",
    "Fy(N)",
    "Fz(N)",
    "Mx(N·m)",
    "My(N·m)",
    "Mz(N·m)",
    "D(N)",
    "L(N)",
    "CD",
    "CY",
    "CL",
    "Cmx",
    "Cmy",
    "Cmz",
    "L/D",
    "Xcp(m)",
    "Xcg(m)",
    "Sref(m2)",
    "Lref(m)",
    "Q(Pa)",
    "Re(1/m)",
    "P(Pa)",
    "T(K)",
]


def write_excel(
    force_report: ForceReport,
    out_path: Union[str, Path],
    sheet_name: Optional[str] = None,
) -> Path:
    """把 ``ForceReport`` 写成 .xlsx 文件,返回输出路径。

    - 单 sheet 输出(``sheet_name`` 默认 "Forces")
    - sheet 名 > 31 字符自动截断
    - 28 列 header(``HEADERS`` 列表),与 ``CoefficientRow`` 字段顺序一一对应
    - 字体 Times New Roman,header 粗体居中,data 行右对齐

    无 ``openpyxl`` 时抛 ``ImportError`` 提示安装 ``inp-tool[post]``。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as e:
        raise ImportError(
            "openpyxl is required for write_excel; "
            "install with: pip install inp-tool[post]"
        ) from e

    out = Path(out_path)
    name = (sheet_name or _DEFAULT_SHEET_NAME)
    if len(name) > _SHEET_NAME_MAX_LEN:
        name = name[:_SHEET_NAME_MAX_LEN]

    wb = Workbook()
    ws = wb.active
    ws.title = name

    font_head = Font(name=_FONT_NAME, bold=True, size=_FONT_SIZE_HEADER)
    font_data = Font(name=_FONT_NAME, size=_FONT_SIZE_DATA)
    align_head = Alignment(horizontal="center", vertical="center")
    align_data = Alignment(horizontal="right", vertical="center")

    # 写 header(第 1 行)
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_head
        cell.alignment = align_head

    # 写 data 行(第 2 行起)
    field_names = [f.name for f in dataclass_fields(CoefficientRow)]
    for row_idx, row in enumerate(force_report.rows, start=2):
        for col_idx, field_name in enumerate(field_names, start=1):
            value = getattr(row, field_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = font_data
            cell.alignment = align_data

    # 列宽自适应
    _autofit_columns(ws)

    wb.save(out)
    return out


def _autofit_columns(ws) -> None:
    """列宽 = min(max_cell_str_len + padding, max_width)。"""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = col_cells[0].column_letter
        for cell in col_cells:
            if cell.value is None:
                continue
            length = len(str(cell.value))
            if length > max_len:
                max_len = length
        ws.column_dimensions[col_letter].width = min(
            max_len + _COL_WIDTH_PADDING, _COL_WIDTH_MAX
        )
